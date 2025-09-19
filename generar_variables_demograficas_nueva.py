from __future__ import annotations
from pathlib import Path
import argparse
import json
import os
import shutil
import sys
import tempfile
import time
from difflib import SequenceMatcher
import unicodedata
import pandas as pd
from openpyxl import load_workbook

# Defaults
INPUT_FILE = 'input.xlsx'
OUTPUT_FILE = 'salida.xlsx'
DEFAULT_SHEET = 'Datos A2'
DEFAULT_SHEET_OUT = 'V. Demogr (nuevo)'
DEFAULT_MAPPING_FILE = 'variables_demograficas_mapping.json'

PROMOTOR_VARIANTS = {'entusiasta','entusiastas','promotor','promotores','promoter','promoters'}
PASIVO_VARIANTS = {'pasivo','pasivos','neutral','neutrales','indiferente','indiferentes'}
DETRACTOR_VARIANTS = {'detractor','detractores'}

# Valores a tratar como vacíos de categoría (no imprimir ni contar en el denominador del grupo)
EMPTY_LIKE = {'', 'nan', 'none', 'na', 'n/a', 'n\\a', '-', '–', '--', 's/d', 'sd'}


def normalize_text(s: str) -> str:
    if s is None:
        return ''
    s = str(s)
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    s = ' '.join(s.replace('\n',' ').replace('\r',' ').replace('\t',' ').split()).strip().lower()
    return s


def order_suffix_map(order_list: list[str]) -> dict[str, str]:
    """Build a normalized mapping from base label (without numeric prefix) to the full ordered label.
    Example: ["0 Femenino", "1 Masculino"] -> {"femenino": "0 Femenino", "masculino": "1 Masculino"}
    Works also for items like "0 Casado (a)" -> key "casado (a)".
    """
    auto = {}
    for item in order_list or []:
        if not item:
            continue
        text = str(item)
        # remove leading numeric prefix like "0 ", "1 ", etc.
        base = text
        # split once on space if starts with digits
        if base and base[0].isdigit():
            parts = base.split(' ', 1)
            if len(parts) == 2:
                base = parts[1]
        auto[normalize_text(base)] = text
    return auto


def open_excel_safely(p: Path) -> pd.ExcelFile:
    try:
        return pd.ExcelFile(p, engine='openpyxl')
    except PermissionError:
        tmpdir = Path(tempfile.gettempdir())
        last_err = None
        # Reintentos con backoff, intentando copiar a temp y leer desde allí
        for i in range(8):
            try:
                tmp_path = tmpdir / f"tmp_{p.stem}_{os.getpid()}_{i}.xlsx"
                shutil.copy2(p, tmp_path)
                return pd.ExcelFile(tmp_path, engine='openpyxl')
            except PermissionError as e:
                last_err = e
                time.sleep(1.0)
        print(f"No se pudo leer '{p}' porque está en uso. Cierra el archivo en Excel/OneDrive y vuelve a intentar, o pasa un archivo alterno con --in-file.", file=sys.stderr)
        raise last_err


def write_excel_safely(out_path: Path, df_outs: dict[str, pd.DataFrame]):
    mode = 'a' if out_path.exists() else 'w'
    writer_kwargs = dict(engine='openpyxl', mode=mode)
    if mode == 'a':
        writer_kwargs['if_sheet_exists'] = 'replace'
    try:
        with pd.ExcelWriter(out_path, **writer_kwargs) as writer:
            for sheet_name, df in df_outs.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
        final_path = out_path
    except PermissionError:
        tmpdir = Path(tempfile.gettempdir())
        temp_out = tmpdir / f"tmp_{out_path.stem}_{os.getpid()}.xlsx"
        with pd.ExcelWriter(temp_out, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in df_outs.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
        try:
            shutil.copy2(temp_out, out_path)
            final_path = out_path
        except PermissionError:
            final_path = temp_out
            print(f'El archivo de salida estaba en uso. Se creó una copia temporal en: {temp_out}', file=sys.stderr)
    return final_path


def remove_sheets_safely(xlsx_path: Path, to_remove: list[str]) -> Path:
    if not to_remove:
        return xlsx_path
    try:
        wb = load_workbook(xlsx_path)
    except PermissionError:
        # copy to temp and operate there
        tmpdir = Path(tempfile.gettempdir())
        temp_copy = tmpdir / f"tmp_rm_{xlsx_path.stem}_{os.getpid()}.xlsx"
        shutil.copy2(xlsx_path, temp_copy)
        wb = load_workbook(temp_copy)
        xlsx_path = temp_copy
    # delete requested sheets if present
    for name in list(to_remove):
        if name in wb.sheetnames and len(wb.sheetnames) > 1:
            ws = wb[name]
            wb.remove(ws)
    # ensure workbook keeps at least one sheet
    if len(wb.sheetnames) == 0:
        wb.create_sheet('Hoja1')
    try:
        wb.save(xlsx_path)
        return xlsx_path
    except PermissionError:
        tmpdir = Path(tempfile.gettempdir())
        temp_out = tmpdir / f"tmp_rm_save_{xlsx_path.stem}_{os.getpid()}.xlsx"
        wb.save(temp_out)
        try:
            shutil.copy2(temp_out, xlsx_path)
            return xlsx_path
        except PermissionError:
            print(f"No se pudo sobrescribir el archivo original; queda una copia con hojas removidas en: {temp_out}", file=sys.stderr)
            return temp_out


def col_letter_to_index(letter: str) -> int:
    if not letter:
        return -1
    s = letter.strip().upper()
    total = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'):
            continue
        total = total * 26 + (ord(ch) - ord('A') + 1)
    return total - 1


def load_mapping(mapping_path: Path) -> dict:
    if not mapping_path.exists():
        raise SystemExit(f'No se encuentra {mapping_path.name} en {mapping_path.parent}')
    with mapping_path.open('r', encoding='utf-8') as f:
        return json.load(f)


def get_series(df: pd.DataFrame, df_raw: pd.DataFrame, letter: str | None) -> pd.Series:
    if not letter:
        return pd.Series([], dtype=object)
    idx = col_letter_to_index(letter)
    if idx < 0 or idx >= df_raw.shape[1]:
        return pd.Series([], dtype=object)
    s = df_raw.iloc[:, idx]
    if len(s) == len(df) + 1:
        s = s.iloc[1:]
    return s.reset_index(drop=True)


def classify_nps(s: pd.Series) -> tuple[pd.Series, pd.Series, pd.Series]:
    sn = s.astype(str).map(normalize_text)
    ent = sn.isin(PROMOTOR_VARIANTS)
    pas = sn.isin(PASIVO_VARIANTS)
    det = sn.isin(DETRACTOR_VARIANTS)
    return det, pas, ent


def generar():
    parser = argparse.ArgumentParser(description='Generar hoja de variables demográficas (conteos y % + NPS) en una hoja nueva.')
    parser.add_argument('--in-file', dest='in_file', help='Archivo de entrada (por defecto input.xlsx en esta carpeta)')
    parser.add_argument('--sheet', help='Hoja de entrada (por defecto Datos A2)')
    parser.add_argument('--out', help='Archivo de salida (por defecto salida.xlsx)')
    parser.add_argument('--sheet-out', dest='sheet_out', help='Nombre de la hoja de salida (por defecto V. Demogr (nuevo))')
    parser.add_argument('--id', dest='ids', nargs='*', help='Filtrar por uno o varios ID (columna A)')
    parser.add_argument('--label-mode', choices=['raw', 'no-prefix', 'prefixed'], default='raw',
                        help='Cómo mostrar la columna "Variable": raw=texto tal cual en la columna, no-prefix=sin prefijo numérico, prefixed=con prefijo (por defecto raw)')
    parser.add_argument('--remove-sheets', nargs='*', help='Nombres de hojas a eliminar del archivo de salida (por ejemplo "V. Demogr")')
    parser.add_argument('--keep-others', action='store_true', help='Conservar otras hojas del libro (por defecto se elimina todo excepto la hoja nueva)')
    parser.add_argument('--mapping-file', dest='mapping_file', help='Ruta a archivo de mapeo JSON (por defecto variables_demograficas_mapping.json)')
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent
    mapping_path = Path(args.mapping_file) if args.mapping_file else (base_dir / DEFAULT_MAPPING_FILE)
    mapping = load_mapping(mapping_path)

    in_path = Path(args.in_file) if args.in_file else (base_dir / INPUT_FILE)
    if not in_path.exists():
        raise SystemExit(f'No se encuentra {INPUT_FILE} en {base_dir}')

    xls = open_excel_safely(in_path)
    hoja = args.sheet or mapping.get('defaults', {}).get('sheet', DEFAULT_SHEET)
    if hoja not in xls.sheet_names:
        wn = hoja.strip().lower()
        best = None
        best_score = 0.0
        for s in xls.sheet_names:
            score = SequenceMatcher(None, wn, s.strip().lower()).ratio()
            if score > best_score:
                best = s
                best_score = score
        hoja = best or xls.sheet_names[0]

    df = xls.parse(hoja)
    df_raw = xls.parse(hoja, header=None)

    # Filtrado por ID si aplica
    id_col = mapping.get('defaults', {}).get('idColumn', 'A')
    if args.ids:
        idx_id = col_letter_to_index(id_col)
        if 0 <= idx_id < df_raw.shape[1]:
            s_id = df_raw.iloc[:, idx_id]
            if len(s_id) == len(df) + 1:
                s_id = s_id.iloc[1:]
            s_id = s_id.reset_index(drop=True).astype(str).str.strip()
            ids_norm = {str(x).strip() for x in args.ids}
            mask = s_id.isin(ids_norm)
            df = df[mask.values].reset_index(drop=True)
            df_raw = df_raw.iloc[1:, :].reset_index(drop=True)

    # Serie NPS clasificada
    nps_letter = mapping.get('defaults', {}).get('npsColumn', 'GV')
    s_nps = get_series(df, df_raw, nps_letter)
    det_mask, pas_mask, ent_mask = classify_nps(s_nps)

    total = len(df)
    # Denominador global para % NPS (todas las filas con NPS no vacías/equivalentes)
    s_nps_norm = s_nps.astype(str).map(normalize_text)
    valid_nps_mask = ~s_nps_norm.isin(EMPTY_LIKE)
    global_nps_total = int(valid_nps_mask.sum())

    rows = []
    index_counter = 0

    for grp in mapping.get('groups', []):
        group_name = grp['group']
        letter = grp.get('column')
        order = grp.get('order', [])
        normalize_map = grp.get('normalize', {})
        s = get_series(df, df_raw, letter)
        if len(order) == 0 and not s.empty:
            order = list(pd.Series(s.astype(str).fillna('')).unique())
        s_raw = s.astype(str).fillna('')
        s_norm = s_raw
        if normalize_map:
            s_norm = s_norm.map(lambda x: normalize_map.get(normalize_text(x), x))
        # Auto-map remaining raw values to the ordered prefixed labels using the order list
        # Only remap if current value is not already one of the ordered labels
        auto_map = order_suffix_map(order)
        if auto_map:
            def _remap_to_order(val: str) -> str:
                if val in order:
                    return val
                return auto_map.get(normalize_text(val), val)
            s_norm = s_norm.map(_remap_to_order)
        # Excluir tokens vacíos/equivalentes del cómputo de categorías
        s_norm_norm = s_norm.astype(str).map(normalize_text)
        valid_mask = ~s_norm_norm.isin(EMPTY_LIKE)
        counts = s_norm[valid_mask].value_counts(dropna=False)
        # Imprimir solo categorías presentes en datos para todas las modalidades
        labels = list(counts.index)
        # Si hay un orden en el JSON, ordenar únicamente las presentes siguiendo ese orden
        if order:
            order_index = {val: i for i, val in enumerate(order)}
            labels.sort(key=lambda x: order_index.get(x, len(order)))
        # Denominador por variable (grupo): suma de sus categorías válidas presentes
        group_total = sum(int(counts.get(lab, 0)) for lab in labels)
        # Pre-calc display text mapping for label-mode=raw (most frequent raw per grouped label)
        display_by_label: dict[str, str] = {}
        if args.label_mode == 'raw' and len(s_norm) > 0:
            df_pair = pd.DataFrame({'raw': s_raw, 'grp': s_norm})
            df_pair_valid = df_pair[valid_mask.values]
            for lab in counts.index:
                subset = df_pair_valid[df_pair_valid['grp'] == lab]['raw']
                if not subset.empty:
                    # most frequent raw string
                    display_by_label[lab] = subset.value_counts(dropna=False).idxmax()
        def strip_prefix_if_any(text: str) -> str:
            t = str(text)
            return t.split(' ', 1)[1] if t and t[0].isdigit() and ' ' in t else t
        for label in labels:
            if pd.isna(label) or str(label).strip() == '':
                continue
            cnt = int(counts.get(label, 0))
            # Porcentaje dentro de la variable (no del total global)
            pct = (cnt / group_total) if group_total else 0.0
            mask_cat = (s_norm == label) & valid_mask
            det = int((det_mask & mask_cat).sum())
            pas = int((pas_mask & mask_cat).sum())
            ent = int((ent_mask & mask_cat).sum())
            # Porcentajes NPS respecto al total global de NPS válidos
            pdetr = (det / global_nps_total) if global_nps_total else 0.0
            ppas = (pas / global_nps_total) if global_nps_total else 0.0
            pent = (ent / global_nps_total) if global_nps_total else 0.0
            # Determine display label according to label-mode
            if args.label_mode == 'prefixed':
                display_label = str(label)
            elif args.label_mode == 'no-prefix':
                display_label = strip_prefix_if_any(label)
            else:  # raw
                if cnt > 0 and label in display_by_label:
                    display_label = str(display_by_label[label])
                else:
                    display_label = strip_prefix_if_any(label)
            rows.append({
                'index': index_counter,
                'Variable': display_label,
                'General': cnt,
                '%': round(pct, 6),
                'Detractores': det,
                'Pasivos': pas,
                'Entusiastas': ent,
                '% Detractores': round(pdetr, 2),
                '% Pasivos': round(ppas, 2),
                '% Entusiastas': round(pent, 2),
                'Grupo': group_name
            })
            index_counter += 1

    df_out = pd.DataFrame(rows, columns=['index','Variable','General','%','Detractores','Pasivos','Entusiastas','% Detractores','% Pasivos','% Entusiastas','Grupo'])

    # Rebalancear la columna '%' para que por cada Grupo sume exactamente 100.00%
    # Se usa asignación por restos mayores sobre 2 decimales en porcentaje (10000 basis points).
    if not df_out.empty:
        for grp_name, idx in df_out.groupby('Grupo').groups.items():
            g = df_out.loc[idx]
            total_cnt = int(g['General'].sum())
            if total_cnt <= 0:
                continue
            # cuotas en centésimas de porcentaje (basis points de %): 100.00% -> 10000
            quotas = (g['General'] * 10000.0) / float(total_cnt)
            floor_bps = quotas.apply(lambda x: int(x // 1))
            remainder = quotas - floor_bps
            need = int(10000 - floor_bps.sum())
            if need > 0:
                # agregar 1 bps a las filas con mayor residuo
                add_index = remainder.sort_values(ascending=False).index[:need]
                floor_bps.loc[add_index] = floor_bps.loc[add_index] + 1
            elif need < 0:
                # en caso raro por redondeos, restar 1 bps a los menores residuos
                sub_index = remainder.sort_values(ascending=True).index[:abs(need)]
                floor_bps.loc[sub_index] = floor_bps.loc[sub_index] - 1
            df_out.loc[idx, '%'] = floor_bps.astype(float) / 10000.0

        # Normalizar las 3 columnas de % NPS por Grupo para que el TOTAL sumado (Detr+Pas+Ent) sea 100.00%
        for grp_name, idx in df_out.groupby('Grupo').groups.items():
            subset = df_out.loc[idx, ['% Detractores','% Pasivos','% Entusiastas']]
            cur_sum = float(subset.to_numpy().sum())
            if cur_sum <= 0:
                df_out.loc[idx, ['% Detractores','% Pasivos','% Entusiastas']] = 0.0
                continue
            # Escalar a suma 1.0 manteniendo proporciones
            scaled = subset / cur_sum
            flat = scaled.stack()
            # Redondeo por restos mayores a 2 decimales en fracción (1.00 -> 100 centésimas)
            quotas_cp = flat * 100.0
            floor_cp = quotas_cp.apply(lambda x: int(x // 1))
            remainder = quotas_cp - floor_cp
            need = int(100 - floor_cp.sum())
            if need > 0:
                add_index = remainder.sort_values(ascending=False).index[:need]
                floor_cp.loc[add_index] = floor_cp.loc[add_index] + 1
            elif need < 0:
                sub_index = remainder.sort_values(ascending=True).index[:abs(need)]
                floor_cp.loc[sub_index] = floor_cp.loc[sub_index] - 1
            adjusted = (floor_cp.astype(float) / 100.0).unstack()
            df_out.loc[idx, ['% Detractores','% Pasivos','% Entusiastas']] = adjusted.values

    out_file = args.out or OUTPUT_FILE
    sheet_out = args.sheet_out or DEFAULT_SHEET_OUT
    final_path = write_excel_safely(base_dir / out_file, {sheet_out: df_out})
    # optionally remove specific sheets
    if args.remove_sheets:
        final_path = remove_sheets_safely(Path(final_path), args.remove_sheets)
    # default behavior: keep only the new sheet unless keep-others is set
    if not args.keep_others:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(final_path)
            for name in list(wb.sheetnames):
                if name != sheet_out and len(wb.sheetnames) > 1:
                    wb.remove(wb[name])
            wb.save(final_path)
        except PermissionError:
            # fall back to temp copy
            final_path = remove_sheets_safely(Path(final_path), [n for n in load_workbook(final_path).sheetnames if n != sheet_out])
    print(f'Hoja "{sheet_out}" escrita en {final_path}')


if __name__ == '__main__':
    generar()
